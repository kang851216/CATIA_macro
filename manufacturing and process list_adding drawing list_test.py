import os
from re import I
import sys
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from datetime import datetime

sys.path.insert(0, os.path.abspath('..\\pycatia'))

from pycatia import catia
from pycatia.enumeration.enumeration_types import cat_work_mode_type

caa = catia()
documents = caa.documents

document = caa.active_document
product = document.product
product.apply_work_mode(cat_work_mode_type.index("DESIGN_MODE"))


class excel:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.create_sheet("开料与加工清单",0)
        self.ds = self.wb.create_sheet("图纸清单",1)
        
        self.ws['A1'].value = "编号"
        #self.ws.merge_cells('B5:G5')
        #self.ws.merge_cells('N5:S5')
        self.ws['B1'].value = "图号"
        self.ws['H1'].value = "类型"
        self.ws['I1'].value = "材质"
        self.ws['J1'].value = "规格"
        self.ws['K1'].value = "（长）"
        self.ws['L1'].value = "（宽）"
        self.ws['M1'].value = "（后）"
        self.ws['N1'].value = "重量(kg)"
        self.ws['O1'].value = "总量"
        self.ws['U1'].value = "加工方式#1"
        self.ws['V1'].value = "加工方式#2"
        self.ws['W1'].value = "备注"
        self.ws.merge_cells('B1:G1')
        self.ws.merge_cells('O1:T1')

        self.ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['H1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['I1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['J1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['K1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['L1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['M1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['N1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['O1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['U1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['V1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['W1'].alignment = Alignment(horizontal="center", vertical="center")
    
        
        self.ws.column_dimensions['A'].width = 5
        self.ws.column_dimensions['B'].width = 10
        self.ws.column_dimensions['C'].width = 10
        self.ws.column_dimensions['D'].width = 10
        self.ws.column_dimensions['E'].width = 10
        self.ws.column_dimensions['F'].width = 10
        self.ws.column_dimensions['G'].width = 10
        self.ws.column_dimensions['H'].width = 9
        self.ws.column_dimensions['I'].width = 12
        self.ws.column_dimensions['J'].width = 30
        self.ws.column_dimensions['K'].width = 7
        self.ws.column_dimensions['L'].width = 7
        self.ws.column_dimensions['M'].width = 7
        self.ws.column_dimensions['N'].width = 10
        self.ws.column_dimensions['O'].width = 3
        self.ws.column_dimensions['P'].width = 3
        self.ws.column_dimensions['Q'].width = 3
        self.ws.column_dimensions['R'].width = 3
        self.ws.column_dimensions['S'].width = 3
        self.ws.column_dimensions['T'].width = 3
        self.ws.column_dimensions['U'].width = 12
        self.ws.column_dimensions['V'].width = 12
        self.ws.column_dimensions['W'].width = 12
        #self.ws.column_dimensions['X'].width = 9
        #self.ws.merge_cells('A1:X2')
        #self.ws['A1'] = "开工与加工清单"
        #self.ws['A3'] = "工号"
        #self.ws['A4'] = "更新日期"
        #self.ws['W3'] = "模型"
        #self.ws['W4'] = "编写人"

        self.ds['A2'].value = "编号"
        self.ds['B2'].value = "类型"
        self.ds['C2'].value = "图号"
        self.ds['D2'].value = "图"
        self.ds['E2'].value = "幅"
        self.ds['F2'].value = "页数"
        self.ds['G2'].value = "版本"
        self.ds['H2'].value = "总生产数量"

        self.ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['C1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['D1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['E1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['F1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['G1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws['H1'].alignment = Alignment(horizontal="center", vertical="center")

        self.ds.column_dimensions['A'].width = 4.5
        self.ds.column_dimensions['B'].width = 8
        self.ds.column_dimensions['C'].width = 15
        self.ds.column_dimensions['D'].width = 3
        self.ds.column_dimensions['E'].width = 3
        self.ds.column_dimensions['F'].width = 4.5
        self.ds.column_dimensions['G'].width = 4.5
        self.ds.column_dimensions['H'].width = 11

    def input(self, input_row, input_column, input_value):
        self.ws.cell(row=input_row, column=input_column).value = input_value

    def save_excel(self):
        self.wb.save("BOM and process list.xlsx")

class process:
    def __init__(self):
        self.iteration = 2
        self.blank = "  "
        self.excel = excel()
        self.fillrowno = 1
        self.partlist = []
    
    def prod_process(self, obje, current_layer, listofpart, newlistofpart):
        if "APC" in obje.part_number:
            listofpart.append(obje.part_number)
            self.iteration += 1
            self.excel.input(self.iteration, 1, self.iteration-1)
            self.excel_write(self.iteration, current_layer, obje)
            self.excel.input(self.iteration, current_layer+13, listofpart.count(obje.part_number))

            self.excel.save_excel()
    
    def quantity_update(self, obje1, current_layer1, listofpart1, newlistofpart1, indexlist):
        self.excel.input(indexlist[newlistofpart1.index(obje1.part_number)], current_layer1+13, listofpart1.count(obje1.part_number))

    def excel_write(self, rowno, columnno, target_obj):
        weight = round(target_obj.analyze.mass,2)
        partno = target_obj.part_number
        definition = target_obj.definition
        self.excel.input(rowno, columnno, partno)
        self.excel.input(rowno, 14, weight)
        category = " "
        definition_text = " "
        if target_obj.is_catpart():
            part_parameters = target_obj.parameters
            
            if part_parameters.is_parameter("Material"):
                materialv = part_parameters.item("Material").value
            if part_parameters.is_parameter("THK"):
                thkv = round(part_parameters.item("THK").value,1)
            if part_parameters.is_parameter("W"):
                Wid = part_parameters.item("W").value
            if part_parameters.is_parameter("L"):
                Len = float(part_parameters.item("L").value)
            if part_parameters.is_parameter("D_in"):
                D_inv = float(part_parameters.item("D_in").value)
            if part_parameters.is_parameter("D_out"):
                Diav = float(part_parameters.item("D_out").value)
            if part_parameters.is_parameter("D"):
                Diav = float(part_parameters.item("D").value)
            if part_parameters.is_parameter("A"):
                Ah = part_parameters.item("A").value
            if part_parameters.is_parameter("B"):
                Bh = part_parameters.item("B").value
            if part_parameters.is_parameter("t"):
                tv = part_parameters.item("t").value
            if part_parameters.is_parameter("model"):
                model = part_parameters.item("model").value
            if part_parameters.is_parameter("Model"):
                model = part_parameters.item("Model").value            

            if part_parameters.is_parameter("W"):
                if part_parameters.is_parameter("L"):
                    if part_parameters.is_parameter("THK"):
                        category = "钢板"
                        definition_text = str(category) + " " + str(int(thkv)) + "THK" + "x" + str(int(Wid))  + "x"+ str(int(Len)) 
            elif part_parameters.is_parameter("D_in"):
                if part_parameters.is_parameter("D_out"):
                    if part_parameters.is_parameter("L"):
                        category = "圆管"
                        definition_text = str(category) + " " + str(int(Diav))  + "x" + str(int(D_inv)) + "x" + "L=" + str(int(Len)) 
                    elif part_parameters.is_parameter("THK"):
                        category = "钢板"
                        definition_text = str(category) + " " + str(int(thkv)) + "THK" + "x" + str(int(Diav))  
            elif part_parameters.is_parameter("D"):
                if part_parameters.is_parameter("THK"):
                        category = "钢板"
                        definition_text = str(category) + " " + str(int(thkv)) + "THK" + "x" + str(int(Diav))              
                elif part_parameters.is_parameter("L"):
                    category = "圆钢"
                    definition_text = str(category) + " " + "D" + str(int(Diav))  + "x" + "L=" + str(int(Len)) 
            elif part_parameters.is_parameter("D_out"): 
                if part_parameters.is_parameter("THK"):
                    category = "钢板"
                    definition_text = str(category) + " " + str(int(thkv)) + "THK" + "x" + str(int(Diav)) 
                elif part_parameters.is_parameter("L"):
                    category = "圆钢"
                    definition_text = str(category) + " " + "D" + str(int(Diav))  + "x" + "L=" + str(int(Len))   
            
            elif part_parameters.is_parameter("A"):
                if part_parameters.is_parameter("t"):
                    if part_parameters.is_parameter("B"):
                        category = "扁通"
                        definition_text = str(model) + "," + "L=" + str(int(Len)) 
                    else:
                        category = "方通"
                        definition_text = str(model) + "," + "L=" + str(int(Len)) 
            
            elif "角钢" in definition:
                category = "角钢"
                if part_parameters.is_parameter("model"):
                    definition_text = str(model) + "," + "L=" + str(int(Len))
                elif part_parameters.is_parameter("Model"):
                    definition_text = str(model) + "," + "L=" + str(int(Len))
            
            elif "槽钢" in definition:
                category = "槽钢"
                if part_parameters.is_parameter("model"):
                    definition_text = str(model) + "," + "L=" + str(int(Len))
                elif part_parameters.is_parameter("Model"):
                    definition_text = str(model) + "," + "L=" + str(int(Len))
            else :
                category = "其他"
                definition_text = target_obj.definition
            
            '''
            elif "扁通" in definition:
                category = "扁通"
                if part_parameters.is_parameter("Model"):
                    definition_text = str(category) + str(model) + "L=" + str(int(Len)) + "mm"
                else:
                    definition_text = target_obj.definition
            elif "圆通" in definition:
                category = "圆通"
                if part_parameters.is_parameter("Model"):
                    definition_text = str(category) + str(model) + "L=" + str(int(Len)) + "mm"
                else:
                    definition_text = target_obj.definition
            elif "方通" in definition:
                category = "方通"
                if part_parameters.is_parameter("Model"):
                    definition_text = str(category) + str(model) + "L=" + str(int(Len)) + "mm"
                else:
                    definition_text = target_obj.definition
            elif "钢板" in definition:
                category = "钣金"
            '''
            
            self.excel.input(rowno, 8, category)
            
            if part_parameters.is_parameter("L"):
                self.excel.input(rowno, 11, Len)
            
            if part_parameters.is_parameter("W"):
                self.excel.input(rowno, 12, Wid)
            
            if part_parameters.is_parameter("THK"): 
                self.excel.input(rowno, 13, thkv)
            elif part_parameters.is_parameter("t"):
                self.excel.input(rowno, 13, tv) 
            
            
            
            if part_parameters.is_parameter("Material"):
                self.excel.input(rowno, 9, materialv)
        
            self.excel.input(rowno, 10, definition_text)
        
        else:
            category = "组装件"
            self.excel.input(rowno, 8, category)
            self.excel.input(rowno, 10, definition_text)
        
p = process()
list_1 = []
newlist_1 = []
pl1 = []
npl1 = []
ql1 = []
index1 = []
stime = datetime.now()
p.excel.input(2,2,product.part_number)
p.excel.input(2,1,1)
for product1 in product.products:
    if "APC" in product1.part_number:
        ql1.append(product1.part_number) 
        if product1.part_number not in pl1:
            npl1.append(product1.part_number)
            p.prod_process(product1, 3, list_1, newlist_1)
            index1.append(p.iteration)
            print("-------------")
            print(index1)
            print(npl1)
            print("-------------")
            if product1.is_catproduct():
                list_2 = []
                newlist_2 = []
                pl2 = []
                npl2 = []
                ql2 = []
                index2 = []
                for product2 in product1.products:
                    if "APC" in product2.part_number:
                        ql2.append(product2.part_number) 
                        if product2.part_number not in pl2:               
                            npl2.append(product2.part_number)
                            p.prod_process(product2, 4, list_2, newlist_2)
                            index2.append(p.iteration)
                            if product2.is_catproduct():
                                list_3 = []
                                newlist_3 = []
                                pl3 = []
                                npl3 = []
                                ql3 = []
                                index3 = []
                                for product3 in product2.products:
                                    if "APC" in product3.part_number:
                                        ql3.append(product3.part_number) 
                                        if product3.part_number not in pl3: 
                                            npl3.append(product3.part_number)
                                            p.prod_process(product3, 5, list_3, newlist_3)
                                            index3.append(p.iteration)
                                            if product3.is_catproduct():
                                                list_4 = []
                                                newlist_4 = []
                                                pl4 = []
                                                npl4 = []
                                                ql4 = []
                                                index4 = []
                                                for product4 in product3.products:
                                                    if "APC" in product4.part_number:
                                                        ql4.append(product4.part_number) 
                                                        if product4.part_number not in pl4: 
                                                            npl4.append(product4.part_number)
                                                            p.prod_process(product4, 6, list_4, newlist_4)
                                                            index4.append(p.iteration)
                                                            if product4.is_catproduct():
                                                                list_5 = []
                                                                newlist_5 = []
                                                                pl5 = []
                                                                npl5 = []
                                                                ql5 = []
                                                                index5 = []
                                                                for product5 in product4.products:
                                                                    if "APC" in product5.part_number:
                                                                        ql5.append(product5.part_number) 
                                                                        if product5.part_number not in pl5: 
                                                                            npl5.append(product5.part_number)
                                                                            p.prod_process(product5, 7, list_5, newlist_5)
                                                                            index5.append(p.iteration)
                                                                            if product5.is_catproduct():
                                                                                list_6 = []
                                                                                newlist_6 = []
                                                                                pl6 = []
                                                                                npl6 = []
                                                                                ql6 = []
                                                                                index6 = []
                                                                                for product6 in product5.products:
                                                                                    if "APC" in product6.part_number:
                                                                                        ql6.append(product6.part_number)
                                                                                        if product6.part_number not in pl6: 
                                                                                            npl6.append(product6.part_number)           
                                                                                            p.prod_process(product6, 8, list_6, newlist_6)
                                                                                            index6.append(p.iteration)             
                                                                                        elif product6.part_number in npl6:
                                                                                            p.quantity_update(product6, 8, ql6, npl6, index6)   
                                                                            #else :
                                                                            #    p.prod_process(product5, 6, list_5, newlist_5)
                                                                            #    index5.append(p.iteration)
                                                                            pl5.append(product5.part_number)
                                                                    
                                                                        elif product4.part_number in npl5:
                                                                            p.quantity_update(product5, 7, ql5, npl5, index5)     
                                                            #else :
                                                            #    p.prod_process(product4, 5, list_4, newlist_4)
                                                            #    index4.append(p.iteration)
                                                            pl4.append(product4.part_number)
                                                        
                                                        elif product4.part_number in npl4:
                                                            p.quantity_update(product4, 6, ql4, npl4, index4)                                           

                                            #else :
                                            #    p.prod_process(product3, 4, list_3, newlist_3)
                                            #    index3.append(p.iteration)
                                            pl3.append(product3.part_number)

                                        elif product3.part_number in npl3:
                                            p.quantity_update(product3, 5, ql3, npl3, index3)       

                            #else :
                            #    p.prod_process(product2, 3, list_2, newlist_2)
                            #    index2.append(p.iteration)
                            pl2.append(product2.part_number)
                        
                        elif product2.part_number in npl2:
                            p.quantity_update(product2, 4, ql2, npl2, index2)                

            #else:
            #    p.prod_process(product1, 2, list_1, newlist_1)
            #    index1.append(p.iteration)
            pl1.append(product1.part_number)
        
        elif product1.part_number in npl1:
            p.quantity_update(product1, 3, ql1, npl1, index1)
p.excel.save_excel()


drawinglist =[]
max = int(p.excel.ws.max_row)
for r in range(2, max):
    for c in range(2, 6):
        drawingno = p.excel.ws.cell(r,c).value
        if drawingno not in drawinglist and drawingno != None:
            drawinglist.append(drawingno)
drawinglist.sort()
for i in range(0,len(drawinglist)):
    p.excel.ds.cell(row=i+3,column=1).value = i+1
    p.excel.ds.cell(row=i+3,column=3).value = drawinglist[i]
    p.excel.ds.cell(row=i+3,column=4).value = "A"
    p.excel.ds.cell(row=i+3,column=5).value = "3"
    p.excel.ds.cell(row=i+3,column=7).value = "A"

qty=0
max = int(p.excel.ds.max_row)
for ii in range(2, max):
    dwgno = str(p.excel.ds.cell(row=ii+1, column=3).value)
    print(dwgno)
    for product1 in product.products:
        if dwgno in product1.part_number:
            qty = qty + 1
        if product1.is_catproduct():
            for product2 in product1.products:
                if dwgno in product2.part_number:
                    qty = qty + 1
                if product2.is_catproduct():
                    for product3 in product2.products:
                        if dwgno in product3.part_number:
                            qty = qty + 1
                        if product3.is_catproduct():
                            for product4 in product3.products:
                                if dwgno in product4.part_number:
                                    qty = qty + 1
                                if product4.is_catproduct():
                                    for product5 in product4.products:
                                        if dwgno in product5.part_number:
                                            qty = qty + 1
                                        if product5.is_catproduct():
                                            for product6 in product5.products:
                                                if dwgno in product6.part_number:
                                                    qty = qty + 1
    p.excel.ds.cell(row=ii+1,column=8).value = qty
    qty=0
p.excel.save_excel()
etime = datetime.now()
print("Start Time: ", stime.strftime("%H:%M:%S"))
print("End Time: ", etime.strftime("%H:%M:%S"))
